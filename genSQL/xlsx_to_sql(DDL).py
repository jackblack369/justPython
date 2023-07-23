import os
import datetime
from openpyxl import load_workbook

path = r'C:\Users\Administrator\Desktop\dkhz.xlsx'
file = open(r'C:\Users\Administrator\Desktop\python\EMP.txt','w')
wb = load_workbook(path)

def len(path,ul):
	i = 0 
	wb = load_workbook(path)
	sheet = wb['{}'.format(ul)]
	for row in sheet.rows:
		i = i + 1
	return i

for sname in wb.sheetnames:
	sheet = wb['{}'.format(sname)]
	file.write('drop table {} cascade constraints;\n\n'.format(sname))

for sname in wb.sheetnames:
	sheet = wb['{}'.format(sname)]
	file.write('create table {}( \n'.format(sname))

	cells = sheet['D2:D{}'.format(len(path,sname))]
	values = []
	for row in cells:
		values.append(row[0].value)

	if values.count(None) == len(path,sname)-1:
		for i in range(1,len(path,sname)+1):
			tit1 = sheet['A{}'.format(i)].value
			tit3 = sheet['C{}'.format(i)].value
			if i == len(path,sname):
				file.write('{}   {}\n'.format(tit1,tit3))

			elif i >= 2:
				file.write('{}   {},\n'.format(tit1,tit3))

		file.write(');\n\n')
	else:
		for i in range(1,len(path,sname)+1):
			tit1 = sheet['A{}'.format(i)].value
			tit3 = sheet['C{}'.format(i)].value
			tit4 = sheet['D{}'.format(i)].value
			
			if i == len(path,sname):
				file.write('{}   {}  '.format(tit1,tit3))
				if tit4 == '主键' or '非空':
					file.write('not null\n')
				else:
					file.write(',\n')
				
			elif i >= 2 and i < len(path,sname) :
				file.write('{}   {}   '.format(tit1,tit3))
				if tit4 == '主键' or '非空':
					file.write('not null,\n')
				else:
					file.write(',\n')
		
		val = 'constraint PK_{} primary key ('.format(sname)
		for i in range(1,len(path,sname)+1):
			tit1 = sheet['A{}'.format(i)].value
			tit4 = sheet['D{}'.format(i)].value
			if tit4 == '主键':
				val = val + '{},'.format(tit1)
		val = val[:-1]
		file.write(val)
		file.write(');\n')

	for i in range(1,len(path,sname)+1):
		tit1 = sheet['A{}'.format(i)].value
		tit2 = sheet['B{}'.format(i)].value
		if i >= 2:
			file.write("comment on column {}.{} is '{}';\n".format(sname,tit1,tit2))

		file.write('\n')